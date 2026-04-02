from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, abort, session, flash
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import sqlite3
import os, datetime, shutil, json, secrets, base64
from openpyxl import Workbook
import csv
from io import StringIO, BytesIO
import qrcode

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "database.db")
BACKUP_DIR = os.path.join(APP_DIR, "backup")

# Public receipt links (WhatsApp): token validity window
PUBLIC_RECEIPT_TOKEN_TTL_DAYS = 7

COLLEGE = {
  "name": "Kaveri University",
  "address": "Gowraram (V), Wargal (M), Siddipet \u2013 502279, Telangana, India.",
  "academic_year": "2025-26",
  "receipt_prefix": "KU/2025-26/RCPT",
  "receipt_start": 1,
  "footer_lines": [
    "This is a computer-generated receipt.",
    "This receipt is valid subject to realization of payment."
  ],
  "signatory_text": "Authorised Signatory (Accounts Section)"
}

# --- PDF Font Setup (supports ₹ symbol) ---
_PDF_FONTS_READY = False

def ensure_pdf_fonts():
  """Register bundled fonts so special symbols like ₹ render correctly in PDFs."""
  global _PDF_FONTS_READY
  if _PDF_FONTS_READY:
    return
  fonts_dir = os.path.join(APP_DIR, 'static', 'fonts')
  regular = os.path.join(fonts_dir, 'DejaVuSans.ttf')
  bold = os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf')
  if os.path.exists(regular) and os.path.exists(bold):
    try:
      pdfmetrics.registerFont(TTFont('KaveriFont', regular))
      pdfmetrics.registerFont(TTFont('KaveriFont-Bold', bold))
      _PDF_FONTS_READY = True
      return
    except Exception:
      pass
  # Fallback to built-in fonts if bundling fails
  _PDF_FONTS_READY = True


def pdf_set_font(p, bold=False, size=10):
  """Use our registered fonts if available; else Helvetica."""
  name = 'KaveriFont-Bold' if bold else 'KaveriFont'
  try:
    p.setFont(name, size)
  except Exception:
    p.setFont('Helvetica-Bold' if bold else 'Helvetica', size)


# --- Fee Heads (Final) ---
# NOTE: Excel column names must match these exactly.
FEE_HEADS = [
  "Tuition Fee",
  "Hostel Fee",
  "Bus Fee",
  "Exam Fee",
  "Uniform Fee",
  "Library Fee",
  "Sports Fee",
  "Lab Fee",
  "Admission Fee",
  "Fine",
  "Other Fee",
]

# fee_profile column mapping (snake_case)
FEE_PROFILE_COLS = {
  "Tuition Fee": "tuition_fee",
  "Hostel Fee": "hostel_fee",
  "Bus Fee": "bus_fee",
  "Exam Fee": "exam_fee",
  "Uniform Fee": "uniform_fee",
  "Library Fee": "library_fee",
  "Sports Fee": "sports_fee",
  "Lab Fee": "lab_fee",
  "Admission Fee": "admission_fee",
  "Fine": "fine_fee",
  "Other Fee": "other_fee",
}
def conn():
  c = sqlite3.connect(DB_PATH)
  c.row_factory = sqlite3.Row
  return c


def get_setting(key, default=None):
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT value FROM settings WHERE key=?", (key,))
  row = cur.fetchone()
  c.close()
  return row["value"] if row else default

def set_setting(key, value):
  c = conn()
  cur = c.cursor()
  cur.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, str(value)))
  c.commit()
  c.close()

def get_current_academic_year():
  return get_setting("current_academic_year", COLLEGE.get("academic_year",""))


def college_context():
  d = dict(COLLEGE)
  d["academic_year"] = get_current_academic_year()
  # Keep receipt prefix in sync with academic year for display (optional)
  try:
    d["receipt_prefix"] = f'KU/{d["academic_year"]}/RCPT'
  except Exception:
    pass
  return d


def ensure_receipts_academic_year_column():
  # Add column if missing (SQLite migration)
  c = conn()
  cur = c.cursor()
  cur.execute("PRAGMA table_info(receipts)")
  cols = [r["name"] for r in cur.fetchall()]
  if "academic_year" not in cols:
    try:
      cur.execute("ALTER TABLE receipts ADD COLUMN academic_year TEXT")
    except Exception:
      pass
  c.commit()
  c.close()

def init_db():
  c = conn()
  cur = c.cursor()

  # --- Master data: Branches & Sections (future-proof folders) ---
  cur.execute("""CREATE TABLE IF NOT EXISTS branches(
    name TEXT PRIMARY KEY,
    full_name TEXT,
    active INTEGER DEFAULT 1
  )""")

  cur.execute("""CREATE TABLE IF NOT EXISTS branch_sections(
    branch_name TEXT NOT NULL,
    section_name TEXT NOT NULL,
    active INTEGER DEFAULT 1,
    PRIMARY KEY(branch_name, section_name),
    FOREIGN KEY(branch_name) REFERENCES branches(name)
  )""")
  cur.execute("""CREATE TABLE IF NOT EXISTS students(
    student_id TEXT PRIMARY KEY,
    admission_number TEXT,
    name TEXT NOT NULL,
    course TEXT,
    branch TEXT,
    year TEXT,
    semester TEXT,
    section TEXT,
    academic_year TEXT,
    whatsapp TEXT,
    email TEXT,
    status TEXT DEFAULT 'ACTIVE'
  )""")

  # Fee profile holds the *planned / total* fee amounts per head for each student.
  # Receipts store actual payments per head.
  cur.execute("""CREATE TABLE IF NOT EXISTS fee_profile(
    student_id TEXT PRIMARY KEY,
    tuition_fee INTEGER DEFAULT 0,
    hostel_fee INTEGER DEFAULT 0,
    bus_fee INTEGER DEFAULT 0,
    exam_fee INTEGER DEFAULT 0,
    uniform_fee INTEGER DEFAULT 0,
    library_fee INTEGER DEFAULT 0,
    sports_fee INTEGER DEFAULT 0,
    lab_fee INTEGER DEFAULT 0,
    admission_fee INTEGER DEFAULT 0,
    fine_fee INTEGER DEFAULT 0,
    other_fee INTEGER DEFAULT 0,
    paid_amount INTEGER DEFAULT 0,
    last_payment_date TEXT,
    FOREIGN KEY(student_id) REFERENCES students(student_id)
  )""")

  cur.execute("""CREATE TABLE IF NOT EXISTS receipts(
    receipt_no TEXT PRIMARY KEY,
    seq INTEGER NOT NULL,
    date TEXT NOT NULL,
    student_id TEXT NOT NULL,
    amount INTEGER NOT NULL,
    mode TEXT NOT NULL,
        transaction_id TEXT,
note TEXT,
    fee_for TEXT, -- JSON array string
    FOREIGN KEY(student_id) REFERENCES students(student_id)
  )""")
  # --- DB migrations (add new columns safely) ---
  def _ensure_column(table, column, coltype):
    cur.execute(f"PRAGMA table_info({table})")
    cols = [r['name'] for r in cur.fetchall()]
    if column not in cols:
      cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {coltype}")

  # academic_year stored per receipt for permanent history (HYBRID design)
  _ensure_column('receipts', 'academic_year', 'TEXT')
  # token to allow public receipt download links (WhatsApp)
  _ensure_column('receipts', 'access_token', 'TEXT')
  _ensure_column('receipts', 'token_issued_at', 'TEXT')
  # older DBs might not have this column in students (safe to ensure)
  _ensure_column('students', 'academic_year', 'TEXT')

  # fee_profile migrations for new fee heads (safe no-op on fresh DBs)
  for _col in [
    'tuition_fee','hostel_fee','bus_fee','exam_fee','uniform_fee','library_fee',
    'sports_fee','lab_fee','admission_fee','fine_fee','other_fee'
  ]:
    _ensure_column('fee_profile', _col, 'INTEGER DEFAULT 0')


  cur.execute("""CREATE TABLE IF NOT EXISTS settings(
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
  )""")

  # --- Auth users ---
  cur.execute("""CREATE TABLE IF NOT EXISTS users(
    username TEXT PRIMARY KEY,
    password_hash TEXT NOT NULL,
    role TEXT NOT NULL DEFAULT 'admin',
    created_at TEXT
  )""")

  # settings init
  # Next receipt sequence
  cur.execute("SELECT value FROM settings WHERE key='next_receipt_seq'")
  row = cur.fetchone()
  if not row:
    cur.execute(
      "INSERT INTO settings(key,value) VALUES('next_receipt_seq', ?)",
      (str(COLLEGE.get("receipt_start", 1000)),)
    )

  # Current academic year (editable in Settings)
  cur.execute("SELECT value FROM settings WHERE key='current_academic_year'")
  row = cur.fetchone()
  if not row:
    cur.execute(
      "INSERT INTO settings(key,value) VALUES('current_academic_year', ?)",
      (COLLEGE.get("academic_year","2025-26"),)
    )

  # Seed default admin user if none exists
  try:
    cur.execute("SELECT COUNT(*) AS n FROM users")
    n = int(cur.fetchone()["n"] or 0)
    if n == 0:
      cur.execute(
        "INSERT INTO users(username, password_hash, role, created_at) VALUES(?,?,?,?)",
        ("admin", generate_password_hash("admin123"), "admin", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
      )
  except Exception:
    pass

  # --- migration: add transaction_id to receipts if missing ---
  try:
    cur.execute("PRAGMA table_info(receipts)")
    cols = [r[1] for r in cur.fetchall()]
    if "transaction_id" not in cols:
      cur.execute("ALTER TABLE receipts ADD COLUMN transaction_id TEXT")
  except Exception:
    pass

  # --- migration: add status to students if missing (for folder / archive system) ---
  try:
    cur.execute("PRAGMA table_info(students)")
    cols = [r[1] for r in cur.fetchall()]
    if "status" not in cols:
      cur.execute("ALTER TABLE students ADD COLUMN status TEXT DEFAULT 'ACTIVE'")
      cur.execute("UPDATE students SET status='ACTIVE' WHERE status IS NULL OR status='' ")
  except Exception:
    pass

  c.commit()

  # Seed default branches/sections if empty (safe for fresh installs)
  try:
    cur.execute("SELECT COUNT(*) AS n FROM branches")
    if int(cur.fetchone()["n"]) == 0:
      cur.execute("INSERT OR IGNORE INTO branches(name, full_name, active) VALUES(?,?,1)", ("CSE", "Computer Science & Engineering"))
      cur.execute("INSERT OR IGNORE INTO branches(name, full_name, active) VALUES(?,?,1)", ("CSE-AIML", "CSE (AI & ML)"))
      for sec in ["A","B","C"]:
        cur.execute("INSERT OR IGNORE INTO branch_sections(branch_name, section_name, active) VALUES(?,?,1)", ("CSE-AIML", sec))
      c.commit()
  except Exception:
    pass

  c.close()


def list_active_branches_and_sections():
  """Return branches and their active sections for sidebar + form options."""
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT name, full_name, active FROM branches ORDER BY name")
  branches = [dict(r) for r in cur.fetchall()]
  cur.execute("SELECT branch_name, section_name, active FROM branch_sections ORDER BY branch_name, section_name")
  sections = [dict(r) for r in cur.fetchall()]
  c.close()

  sec_map = {}
  for s in sections:
    sec_map.setdefault(s["branch_name"], []).append(s)
  for b in branches:
    b["sections"] = sec_map.get(b["name"], [])
  return branches

def seed_dummy():
  # Dummy data disabled (admin will add students manually)
  return


def fee_totals(fp):
  heads = list(FEE_PROFILE_COLS.values())
  total = sum(int(fp.get(h) or 0) if isinstance(fp, dict) else int(fp[h] or 0) for h in heads)
  paid = int((fp.get("paid_amount") if isinstance(fp, dict) else fp["paid_amount"]) or 0)
  due = max(0, total - paid)
  status = "Paid" if due == 0 else ("Pending" if paid==0 else "Partial")
  return total, paid, due, status


def _parse_receipt_items(fee_for_raw):
  """Backward compatible.

  Old DB: fee_for was JSON list of strings.
  New DB: fee_for is JSON dict: {"items": [{"head":...,"amount":...,"label":...}]}
  """
  if not fee_for_raw:
    return []
  try:
    obj = json.loads(fee_for_raw)
  except Exception:
    return []
  if isinstance(obj, list):
    # old style: list of heads, amounts unknown -> treat whole as one line item
    return [{"head": ", ".join([str(x) for x in obj if x]), "amount": None, "label": None}]
  if isinstance(obj, dict):
    items = obj.get("items")
    if isinstance(items, list):
      out = []
      for it in items:
        if not isinstance(it, dict):
          continue
        head = (it.get("head") or "").strip()
        amt = it.get("amount")
        lbl = (it.get("label") or "").strip() or None
        if head:
          out.append({"head": head, "amount": amt, "label": lbl})
      return out
  return []


def payments_by_head(student_id: str):
  """Sum paid amounts per fee head from receipts for a student."""
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT fee_for, amount FROM receipts WHERE student_id=?", (student_id,))
  rows = cur.fetchall()
  c.close()
  sums = {h: 0 for h in FEE_HEADS}
  for r in rows:
    items = _parse_receipt_items(r["fee_for"])
    if items and all(it.get("amount") is not None for it in items):
      for it in items:
        head = it.get("head")
        if head == "Other Fee" and it.get("label"):
          # label is for printing only; head remains Other Fee for accounting
          head = "Other Fee"
        if head in sums:
          try:
            sums[head] += int(it.get("amount") or 0)
          except Exception:
            pass
    else:
      # old receipts: allocate entire amount across the first head if possible
      try:
        amt = int(r["amount"] or 0)
      except Exception:
        amt = 0
      if items:
        head_guess = items[0].get("head")
        if head_guess in sums:
          sums[head_guess] += amt
        else:
          # fallback bucket
          sums.setdefault("Other Fee", 0)
          sums["Other Fee"] += amt
  return sums


def headwise_due(fp_row, paid_sums: dict):
  """Return head-wise due amounts based on fee_profile totals - paid_sums."""
  dues = {}
  for head, col in FEE_PROFILE_COLS.items():
    try:
      planned = int(fp_row[col] or 0)
    except Exception:
      planned = 0
    paid = int(paid_sums.get(head, 0) or 0)
    due = planned - paid
    if due > 0:
      dues[head] = due
  return dues

def get_student_bundle(student_id):
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT * FROM students WHERE student_id=?", (student_id,))
  s = cur.fetchone()
  if not s:
    c.close()
    return None
  cur.execute("SELECT * FROM fee_profile WHERE student_id=?", (student_id,))
  fp = cur.fetchone()
  if not fp:
    # ensure fee profile exists
    cur.execute("INSERT INTO fee_profile(student_id) VALUES(?)", (student_id,))
    c.commit()
    cur.execute("SELECT * FROM fee_profile WHERE student_id=?", (student_id,))
    fp = cur.fetchone()
  total, paid, due, status = fee_totals(fp)
  c.close()
  return {"student": dict(s), "fee": dict(fp), "calc": {"total": total, "paid": paid, "due": due, "status": status}}

def next_receipt_seq(cur):
  cur.execute("SELECT value FROM settings WHERE key='next_receipt_seq'")
  seq = int(cur.fetchone()["value"])
  return seq

def bump_receipt_seq(cur, new_seq):
  cur.execute("UPDATE settings SET value=? WHERE key='next_receipt_seq'", (str(new_seq),))

def make_receipt_no(seq:int):
  return f'{COLLEGE["receipt_prefix"]}/{seq:04d}'


def _parse_dt_any(s: str):
  if not s:
    return None
  try:
    # ISO datetime
    return datetime.datetime.fromisoformat(s)
  except Exception:
    pass
  try:
    # date only
    return datetime.datetime.strptime(s, "%Y-%m-%d")
  except Exception:
    return None

def _token_not_expired(token_issued_at: str, receipt_date: str) -> bool:
  # If parsing fails, allow (fail-open) to avoid blocking legitimate downloads
  dt = _parse_dt_any(token_issued_at) or _parse_dt_any(receipt_date)
  if not dt:
    return True
  delta = datetime.datetime.now() - dt
  return delta.days <= PUBLIC_RECEIPT_TOKEN_TTL_DAYS


def _row_get(row, key: str, default=None):
  """Safe getter for sqlite3.Row or dict."""
  if row is None:
    return default
  if isinstance(row, dict):
    return row.get(key, default)
  try:
    return row[key]
  except Exception:
    return default

def _is_valid_public_receipt_token_seq(seq: int, token: str) -> bool:
  """Allow public receipt links without login using per-receipt token."""
  if not token:
    return False
  try:
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT access_token, token_issued_at, date FROM receipts WHERE seq=?", (seq,))
    row = cur.fetchone()
    c.close()
    if not row:
      return False
    return (str(_row_get(row, "access_token", "") or "") == str(token)) and _token_not_expired(
      _row_get(row, "token_issued_at", "") or "",
      _row_get(row, "date", "") or "",
    )
  except Exception:
    return False

def _is_valid_public_receipt_token_rno(receipt_no: str, token: str) -> bool:
  if not token:
    return False
  try:
    c = conn()
    cur = c.cursor()
    cur.execute("SELECT access_token, token_issued_at, date FROM receipts WHERE receipt_no=?", (receipt_no,))
    row = cur.fetchone()
    c.close()
    if not row:
      return False
    return (str(_row_get(row, "access_token", "") or "") == str(token)) and _token_not_expired(
      _row_get(row, "token_issued_at", "") or "",
      _row_get(row, "date", "") or "",
    )
  except Exception:
    return False

def _public_receipt_request_allowed() -> bool:
  """True if request is a receipt view/pdf with valid token."""
  if session.get("user"):
    return True
  p = request.path or ""
  token = request.args.get("t") or request.args.get("token") or ""
  # Short links
  if p.startswith("/r/"):
    parts = [x for x in p.split("/") if x]
    if len(parts) < 2:
      return False
    try:
      seq = int(parts[1])
    except Exception:
      return False
    return _is_valid_public_receipt_token_seq(seq, token)

  # Full receipt links
  if p.startswith("/receipt/"):
    rest = p[len("/receipt/"):]
    if rest.endswith("/pdf"):
      receipt_no = rest[:-4]
    else:
      receipt_no = rest
    return _is_valid_public_receipt_token_rno(receipt_no, token)

  return False


app = Flask(__name__)
# Secret key for sessions (change in production)
app.secret_key = os.environ.get("KU_ERP_SECRET_KEY", "KU_ERP_CHANGE_ME_2026")

def login_required(fn):
  @wraps(fn)
  def wrapper(*args, **kwargs):
    if not session.get("user"):
      # Allow public receipt links via token (WhatsApp)
      if _public_receipt_request_allowed():
        return fn(*args, **kwargs)
      if request.path.startswith("/api/") or request.path.startswith("/export/") or request.path.startswith("/import/"):
        return jsonify({"ok": False, "error": "AUTH_REQUIRED"}), 401
      return redirect(url_for("login", next=request.path))
    return fn(*args, **kwargs)
  return wrapper


def roles_required(*roles):
  def decorator(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
      if not session.get("user"):
        # Let login_required handle redirects/json errors
        return login_required(fn)(*args, **kwargs)
      current_role = (session.get("role") or "").lower()
      if current_role not in [r.lower() for r in roles]:
        # JSON for API endpoints; 403 for pages
        if request.path.startswith("/api/") or request.path.startswith("/import/"):
          return jsonify({"ok": False, "error": "FORBIDDEN"}), 403
        abort(403)
      return fn(*args, **kwargs)
    return wrapper
  return decorator

def admin_required(fn):
  return roles_required("admin")(fn)

@app.before_request
def _auth_guard():
  if request.endpoint in ("login", "logout", "static"):
    return
  if request.path.startswith("/static/"):
    return
  # Public receipt routes should not be forced to login; routes themselves enforce token
  if request.path.startswith("/r/") or (request.path.startswith("/receipt/") and request.path.endswith("/pdf")):
    return
  if not session.get("user"):
    # Allow public receipt links via token (WhatsApp)
    if _public_receipt_request_allowed():
      return
    if request.path.startswith("/api/") or request.path.startswith("/export/") or request.path.startswith("/import/"):
      return jsonify({"ok": False, "error": "AUTH_REQUIRED"}), 401
    return redirect(url_for("login", next=request.path))

def qr_data_uri(data: str, box_size: int = 6, border: int = 2) -> str:
  """Return a PNG QR code as a data URI."""
  try:
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=box_size, border=border)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return "data:image/png;base64," + b64
  except Exception:
    return ""

@app.route("/login", methods=["GET", "POST"])
def login():
  next_url = request.args.get("next") or url_for("dashboard")
  if request.method == "POST":
    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    if not username or not password:
      flash("Please enter username and password.", "danger")
      return render_template("login.html", college=college_context(), next_url=next_url)

    c = conn()
    cur = c.cursor()
    cur.execute("SELECT username, password_hash, role FROM users WHERE username=?", (username,))
    u = cur.fetchone()
    c.close()

    if u and check_password_hash(u["password_hash"], password):
      session["user"] = u["username"]
      session["role"] = u["role"]
      return redirect(next_url)
    flash("Invalid username or password.", "danger")
  return render_template("login.html", college=college_context(), next_url=next_url)

@app.route("/logout")
def logout():
  session.clear()
  return redirect(url_for("login"))

def _startup():
  init_db()
@app.route("/")
@login_required
def dashboard():
  # Dashboard summary (overall) + Today metrics
  sel_branch = (request.args.get("branch", "ALL") or "ALL").strip()
  sel_section = (request.args.get("section", "ALL") or "ALL").strip()

  c = conn()
  cur = c.cursor()

  # Overall counts
  cur.execute("SELECT COUNT(*) AS n FROM students")
  students = int(cur.fetchone()["n"] or 0)
  cur.execute("SELECT COUNT(*) AS n FROM receipts")
  receipts = int(cur.fetchone()["n"] or 0)

  # Overall due count & totals
  cur.execute("SELECT * FROM fee_profile")
  rows = cur.fetchall()
  due_count = 0
  due_total = 0
  collected_total = 0
  for r in rows:
    total, paid, due, status = fee_totals(r)
    if due > 0:
      due_count += 1
      due_total += due
    collected_total += paid

  # Branch/Section options (dynamic)
  cur.execute("SELECT name, full_name, active FROM branches ORDER BY name")
  branch_rows = [dict(r) for r in cur.fetchall()]
  branches = [b for b in branch_rows if int(b.get("active", 1)) == 1]

  sections_by_branch = {}
  cur.execute("SELECT branch_name, section_name, active FROM branch_sections ORDER BY branch_name, section_name")
  for r in cur.fetchall():
    if int(r["active"] or 0) != 1:
      continue
    sections_by_branch.setdefault((r["branch_name"] or "").strip(), []).append((r["section_name"] or "").strip())

  # Flat list of sections (for template fallback)
  all_sections = sorted({sec for secs in sections_by_branch.values() for sec in secs if sec})


  # Today metrics (optionally filtered by branch/section)
  today = datetime.date.today().isoformat()
  where = ""
  params = [today]
  if sel_branch and sel_branch != "ALL":
    where += " AND s.branch=?"
    params.append(sel_branch)
  if sel_section and sel_section != "ALL":
    where += " AND s.section=?"
    params.append(sel_section)

  cur.execute(
    "SELECT COALESCE(SUM(r.amount),0) AS total "
    "FROM receipts r JOIN students s ON s.student_id=r.student_id "
    "WHERE r.date=?" + where + " AND (s.status IS NULL OR s.status='ACTIVE')",
    params,
  )
  today_total = int(cur.fetchone()["total"] or 0)

  cur.execute(
    "SELECT COUNT(DISTINCT r.student_id) AS n "
    "FROM receipts r JOIN students s ON s.student_id=r.student_id "
    "WHERE r.date=?" + where + " AND (s.status IS NULL OR s.status='ACTIVE')",
    params,
  )
  paid_today = int(cur.fetchone()["n"] or 0)

  # Total active students for same filter
  w = "WHERE (status IS NULL OR status='ACTIVE')"
  params2 = []
  if sel_branch and sel_branch != "ALL":
    w += " AND branch=?"
    params2.append(sel_branch)
  if sel_section and sel_section != "ALL":
    w += " AND section=?"
    params2.append(sel_section)
  cur.execute("SELECT COUNT(*) AS n FROM students " + w, params2)
  total_active = int(cur.fetchone()["n"] or 0)

  not_paid_today = max(total_active - paid_today, 0)

  c.close()

  return render_template(
    "dashboard.html",
    college=college_context(),
    stats={
      "students": students,
      "receipts": receipts,
      "due_count": due_count,
      "due_total": due_total,
      "collected_total": collected_total,
      "today": today,
      "today_total": today_total,
      "paid_today": paid_today,
      "not_paid_today": not_paid_today,
      "total_active": total_active,
    },
    branches=branches,
    sections_by_branch=sections_by_branch,
    all_sections=all_sections,
    sel_branch=sel_branch,
    sel_section=sel_section,
  )

@app.route("/students")
@login_required
def students_page():
  # Folder-based view (branch / section wise) + archive
  q = request.args.get("q", "").strip()
  view = (request.args.get("view", "active") or "active").lower().strip()
  branch = (request.args.get("branch", "") or "").strip()
  section = (request.args.get("section", "") or "").strip()

  # Normalize
  q_like = f"%{q.strip()}%"
  is_removed = (view == "removed")
  is_unassigned = (view == "unassigned")

  c = conn()
  cur = c.cursor()

  # Load active branches & sections for dynamic folder sidebar
  cur.execute("SELECT name, full_name, active FROM branches ORDER BY name")
  branch_rows = [dict(r) for r in cur.fetchall()]
  active_branches = [b for b in branch_rows if int(b.get("active",1)) == 1]

  sections_by_branch = {}
  cur.execute("SELECT branch_name, section_name, active FROM branch_sections ORDER BY section_name")
  for r in cur.fetchall():
    br = (r["branch_name"] or "").strip()
    if not br:
      continue
    sections_by_branch.setdefault(br, []).append({
      "name": (r["section_name"] or "").strip(),
      "active": int(r["active"] or 0)
    })

  folder_tree = []
  for b in active_branches:
    bname = b["name"]
    secs = [s for s in sections_by_branch.get(bname, []) if s["active"] == 1 and s["name"]]
    folder_tree.append({
      "name": bname,
      "full_name": b.get("full_name") or "",
      "sections": secs
    })

  # Sidebar counts (fixed folders + dynamic counts)
  def _count(where_sql="", params=()):
    cur.execute(f"SELECT COUNT(*) AS n FROM students {where_sql}", params)
    return int(cur.fetchone()["n"])

  counts = {
    "all_active": _count("WHERE IFNULL(status,'ACTIVE')='ACTIVE'"),
    "removed": _count("WHERE IFNULL(status,'ACTIVE')='REMOVED'"),
    "unassigned": _count("WHERE IFNULL(status,'ACTIVE')='ACTIVE' AND (branch IS NULL OR TRIM(branch)='')"),
  }

  # dynamic counts for each branch/section
  for node in folder_tree:
    bname = node["name"].upper()
    counts[node["name"]] = _count("WHERE IFNULL(status,'ACTIVE')='ACTIVE' AND UPPER(TRIM(branch))=?", (bname,))
    for sec in node.get("sections", []):
      sname = sec["name"].upper()
      counts[f"{node['name']}:{sec['name']}"] = _count(
        "WHERE IFNULL(status,'ACTIVE')='ACTIVE' AND UPPER(TRIM(branch))=? AND UPPER(TRIM(section))=?",
        (bname, sname)
      )

  # Main students query
  where = []
  params = []

  if is_removed:
    where.append("IFNULL(status,'ACTIVE')='REMOVED'")
  else:
    where.append("IFNULL(status,'ACTIVE')='ACTIVE'")

  if is_unassigned:
    where.append("(branch IS NULL OR TRIM(branch)='')")
  else:
    if branch:
      where.append("UPPER(TRIM(branch))=?")
      params.append(branch.upper())
    if section:
      where.append("UPPER(TRIM(section))=?")
      params.append(section.upper())

  if q:
    where.append("(student_id LIKE ? OR name LIKE ? OR admission_number LIKE ?)")
    params.extend([q_like, q_like, q_like])

  where_sql = "WHERE " + " AND ".join(where) if where else ""
  cur.execute(f"SELECT * FROM students {where_sql} ORDER BY student_id LIMIT 500", tuple(params))
  students = [dict(r) for r in cur.fetchall()]
  c.close()

  return render_template(
    "students.html",
    college=college_context(),
    students=students,
    q=q,
    view=view,
    branch=branch,
    section=section,
    counts=counts,
    folder_tree=folder_tree,
    branches=[n["name"] for n in folder_tree],
  )


@app.route("/branches")
@login_required
@admin_required
def branches_page():
  """Admin page to add/enable/disable branches and sections (non-deletable)."""
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT name, full_name, active FROM branches ORDER BY name")
  branches = [dict(r) for r in cur.fetchall()]
  cur.execute("SELECT branch_name, section_name, active FROM branch_sections ORDER BY branch_name, section_name")
  sections = [dict(r) for r in cur.fetchall()]
  secs_by_branch = {}
  for s in sections:
    secs_by_branch.setdefault(s["branch_name"], []).append(s)
  c.close()
  return render_template("branches.html", college=college_context(), branches=branches, secs_by_branch=secs_by_branch)


@app.route("/branches/add", methods=["POST"])
@login_required
@admin_required
def branches_add():
  name = (request.form.get("name") or "").strip()
  full = (request.form.get("full_name") or "").strip()
  if not name:
    return redirect(url_for("branches_page"))
  c = conn()
  cur = c.cursor()
  cur.execute("INSERT OR IGNORE INTO branches(name, full_name, active) VALUES(?,?,1)", (name, full))
  # if already exists, update full_name if provided
  if full:
    cur.execute("UPDATE branches SET full_name=? WHERE name=?", (full, name))
  c.commit()
  c.close()
  return redirect(url_for("branches_page"))


@app.route("/branches/toggle", methods=["POST"])
@login_required
@admin_required
def branches_toggle():
  name = (request.form.get("name") or "").strip()
  if not name:
    return redirect(url_for("branches_page"))
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT active FROM branches WHERE name=?", (name,))
  row = cur.fetchone()
  if row is not None:
    new_active = 0 if int(row["active"] or 0) == 1 else 1
    cur.execute("UPDATE branches SET active=? WHERE name=?", (new_active, name))
    # when disabling a branch, also disable its sections (keeps UI clean)
    if new_active == 0:
      cur.execute("UPDATE branch_sections SET active=0 WHERE branch_name=?", (name,))
  c.commit()
  c.close()
  return redirect(url_for("branches_page"))


@app.route("/sections/add", methods=["POST"])
@login_required
@admin_required
def sections_add():
  branch_name = (request.form.get("branch_name") or "").strip()
  section_name = (request.form.get("section_name") or "").strip()
  if not branch_name or not section_name:
    return redirect(url_for("branches_page"))
  c = conn()
  cur = c.cursor()
  # ensure branch exists
  cur.execute("INSERT OR IGNORE INTO branches(name, full_name, active) VALUES(?, '', 1)", (branch_name,))
  cur.execute("INSERT OR IGNORE INTO branch_sections(branch_name, section_name, active) VALUES(?,?,1)", (branch_name, section_name))
  c.commit()
  c.close()
  return redirect(url_for("branches_page"))


@app.route("/sections/toggle", methods=["POST"])
@login_required
@admin_required
def sections_toggle():
  branch_name = (request.form.get("branch_name") or "").strip()
  section_name = (request.form.get("section_name") or "").strip()
  if not branch_name or not section_name:
    return redirect(url_for("branches_page"))
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT active FROM branch_sections WHERE branch_name=? AND section_name=?", (branch_name, section_name))
  row = cur.fetchone()
  if row is not None:
    new_active = 0 if int(row["active"] or 0) == 1 else 1
    cur.execute("UPDATE branch_sections SET active=? WHERE branch_name=? AND section_name=?", (new_active, branch_name, section_name))
  c.commit()
  c.close()
  return redirect(url_for("branches_page"))

@app.route("/api/student/<student_id>")
@login_required
def api_student(student_id):
  bundle = get_student_bundle(student_id.strip().upper())
  if not bundle:
    return jsonify({"ok": False, "error": "Student not found"}), 404
  return jsonify({"ok": True, **bundle})

@app.route("/api/student/save", methods=["POST"])
@login_required
def api_student_save():
  data = request.get_json(force=True)
  sid = (data.get("student_id","") or "").strip().upper()
  if not sid:
    return jsonify({"ok": False, "error": "Student ID required"}), 400
  name = (data.get("name","") or "").strip()
  if not name:
    return jsonify({"ok": False, "error": "Name required"}), 400

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT 1 FROM students WHERE student_id=?", (sid,))
  exists = cur.fetchone() is not None

  fields = ("admission_number","course","branch","year","semester","section","academic_year","whatsapp","email")
  values = [data.get(f,"") for f in fields]

  if not exists:
    cur.execute("""INSERT INTO students(student_id, admission_number, name, course, branch, year, semester, section, academic_year, whatsapp, email, status)
                  VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (sid, values[0], name, values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8], 'ACTIVE'))
    cur.execute("INSERT OR IGNORE INTO fee_profile(student_id) VALUES(?)", (sid,))
  else:
    cur.execute("""UPDATE students SET admission_number=?, name=?, course=?, branch=?, year=?, semester=?, section=?, academic_year=?, whatsapp=?, email=?
                  WHERE student_id=?""",
                (values[0], name, values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8], sid))

  # fee profile update
  fee_fields = (
    "tuition_fee","hostel_fee","bus_fee","exam_fee","uniform_fee","library_fee","sports_fee","lab_fee",
    "admission_fee","fine_fee","other_fee",
    "paid_amount","last_payment_date"
  )
  fee_vals = []
  for f in fee_fields:
    v = data.get(f, 0 if f!="last_payment_date" else None)
    fee_vals.append(v)
  cur.execute("""UPDATE fee_profile SET tuition_fee=?, hostel_fee=?, bus_fee=?, exam_fee=?, uniform_fee=?, library_fee=?, sports_fee=?, lab_fee=?,
                                      admission_fee=?, fine_fee=?, other_fee=?,
                                      paid_amount=?, last_payment_date=?
                 WHERE student_id=?""",
              (*fee_vals, sid))

  c.commit()
  c.close()
  return jsonify({"ok": True})

@app.route("/api/student/delete/<student_id>", methods=["POST"])
@login_required
@admin_required
def api_student_delete(student_id):
  sid = (student_id or "").strip().upper()
  if not sid:
    return jsonify({"ok": False, "error": "Student ID required"}), 400

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT 1 FROM students WHERE student_id=?", (sid,))
  if cur.fetchone() is None:
    c.close()
    return jsonify({"ok": False, "error": "Student not found"}), 404

  # Soft remove (archive) so receipts/history are kept safely.
  cur.execute("UPDATE students SET status='REMOVED' WHERE student_id=?", (sid,))
  removed = cur.rowcount

  c.commit()
  c.close()
  return jsonify({"ok": True, "removed": removed})


@app.route("/api/student/restore/<student_id>", methods=["POST"])
@login_required
@admin_required
def api_student_restore(student_id):
  sid = (student_id or "").strip().upper()
  if not sid:
    return jsonify({"ok": False, "error": "Student ID required"}), 400

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT 1 FROM students WHERE student_id=?", (sid,))
  if cur.fetchone() is None:
    c.close()
    return jsonify({"ok": False, "error": "Student not found"}), 404

  cur.execute("UPDATE students SET status='ACTIVE' WHERE student_id=?", (sid,))
  restored = cur.rowcount
  c.commit()
  c.close()
  return jsonify({"ok": True, "restored": restored})


@app.route("/api/student/purge/<student_id>", methods=["POST"])
@login_required
@admin_required
def api_student_purge(student_id):
  """Permanent delete (Admin-only by convention)."""
  sid = (student_id or "").strip().upper()
  if not sid:
    return jsonify({"ok": False, "error": "Student ID required"}), 400

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT 1 FROM students WHERE student_id=?", (sid,))
  if cur.fetchone() is None:
    c.close()
    return jsonify({"ok": False, "error": "Student not found"}), 404

  cur.execute("DELETE FROM receipts WHERE student_id=?", (sid,))
  receipts_deleted = cur.rowcount
  cur.execute("DELETE FROM fee_profile WHERE student_id=?", (sid,))
  fee_deleted = cur.rowcount
  cur.execute("DELETE FROM students WHERE student_id=?", (sid,))
  student_deleted = cur.rowcount
  c.commit()
  c.close()
  return jsonify({"ok": True, "deleted": {"students": student_deleted, "fee_profile": fee_deleted, "receipts": receipts_deleted}})


@app.route("/payments")
@login_required
def payments_page():
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT * FROM receipts ORDER BY seq DESC LIMIT 50")
  receipts = [dict(r) for r in cur.fetchall()]
  # expand fee_for (new dict items + backward compatible list)
  for r in receipts:
    # sqlite3.Row supports mapping access (r["col"]) but not .get()
    fee_for_raw = r["fee_for"] if (hasattr(r, "keys") and ("fee_for" in r.keys())) else None
    items = _parse_receipt_items(fee_for_raw or "")
    if items and items[0].get("amount") is None:
      # old style
      r["fee_for_list"] = [items[0].get("head")]
    else:
      out = []
      for it in items:
        if it.get("head") == "Other Fee" and it.get("label"):
          out.append(it.get("label"))
        else:
          out.append(it.get("head"))
      r["fee_for_list"] = out
  c.close()
  return render_template("payments.html", college=college_context(), fee_heads=FEE_HEADS, receipts=receipts)

@app.route("/api/receipt/create", methods=["POST"])
@login_required
def api_receipt_create():
  data = request.get_json(force=True)
  sid = (data.get("student_id","") or "").strip().upper()
  # New: item-wise amounts per selected fee head
  items = data.get("items")
  fee_for = data.get("fee_for") or []
  amount = int(data.get("amount") or 0)
  mode = (data.get("mode","") or "").strip()
  note = (data.get("note","") or "").strip()
  transaction_id = (data.get("transaction_id","") or "").strip()
  # Normalize items -> store in receipts.fee_for as JSON dict
  normalized_items = []
  if isinstance(items, list) and len(items) > 0:
    for it in items:
      if not isinstance(it, dict):
        continue
      head = (it.get("head") or "").strip()
      if head not in FEE_HEADS:
        continue
      try:
        amt = int(it.get("amount") or 0)
      except Exception:
        amt = 0
      if amt <= 0:
        continue
      label = (it.get("label") or "").strip()
      if head == "Other Fee":
        if not label:
          return jsonify({"ok": False, "error": "Enter Other Fee name"}), 400
      normalized_items.append({"head": head, "amount": amt, "label": label or None})

    amount = sum(int(x["amount"]) for x in normalized_items)
  else:
    # Backward compatible: old style fee_for list + single amount
    if not isinstance(fee_for, list):
      fee_for = []
    fee_for = [str(x) for x in fee_for if x]
    normalized_items = []
  if not sid:
    return jsonify({"ok": False, "error": "Student ID required"}), 400
  if amount <= 0:
    return jsonify({"ok": False, "error": "Amount must be > 0"}), 400
  if not mode:
    return jsonify({"ok": False, "error": "Payment mode required"}), 400
  if mode.lower() in ["upi","bank transfer","bank","neft","rtgs","imps"] and not transaction_id:
    return jsonify({"ok": False, "error": "Transaction ID / UTR No is required for UPI/Bank payments"}), 400
  if len(normalized_items) == 0 and (not isinstance(fee_for, list) or len(fee_for) == 0):
    return jsonify({"ok": False, "error": "Select at least one fee head"}), 400

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT * FROM students WHERE student_id=?", (sid,))
  s = cur.fetchone()
  if not s:
    c.close()
    return jsonify({"ok": False, "error": "Student not found"}), 404

  seq = next_receipt_seq(cur)
  rno = make_receipt_no(seq)
  today = datetime.date.today().isoformat()
  access_token = secrets.token_urlsafe(16)
  token_issued_at = datetime.datetime.now().isoformat(timespec='seconds')

  fee_for_payload = {"items": normalized_items} if len(normalized_items) > 0 else fee_for
  cur.execute("""INSERT INTO receipts(receipt_no, seq, date, student_id, amount, mode, transaction_id, note, fee_for, academic_year, access_token, token_issued_at)
                 VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
              (rno, seq, today, sid, amount, mode, transaction_id, note, json.dumps(fee_for_payload), get_current_academic_year(), access_token, token_issued_at))

  # update fee_profile paid and last_payment_date
  cur.execute("SELECT * FROM fee_profile WHERE student_id=?", (sid,))
  fp = cur.fetchone()
  paid = int(fp["paid_amount"] or 0) + amount
  cur.execute("UPDATE fee_profile SET paid_amount=?, last_payment_date=? WHERE student_id=?", (paid, today, sid))

  bump_receipt_seq(cur, seq+1)
  c.commit()
  c.close()
  return jsonify({"ok": True, "receipt_no": rno})

@app.route("/receipt/<path:receipt_no>")
@login_required
def receipt_view(receipt_no):
  # receipt_no includes slashes; Flask path captures it
  rn = receipt_no
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT * FROM receipts WHERE receipt_no=?", (rn,))
  r = cur.fetchone()
  if not r:
    c.close()
    abort(404)

  # Ensure access_token exists (for public WhatsApp links) and set issued timestamp
  if not (r.get("access_token") if isinstance(r, dict) else r["access_token"]):
    try:
      new_tok = secrets.token_urlsafe(16)
      issued = datetime.datetime.now().isoformat(timespec="seconds")
      cur.execute("UPDATE receipts SET access_token=?, token_issued_at=? WHERE receipt_no=?", (new_tok, issued, rn))
      c.commit()
      cur.execute("SELECT * FROM receipts WHERE receipt_no=?", (rn,))
      r = cur.fetchone()
    except Exception:
      pass

  # For older rows, ensure token_issued_at exists (used for expiry check)
  try:
    if not (r.get("token_issued_at") if isinstance(r, dict) else r["token_issued_at"]):
      issued = datetime.datetime.now().isoformat(timespec="seconds")
      cur.execute("UPDATE receipts SET token_issued_at=? WHERE receipt_no=?", (issued, rn))
      c.commit()
      cur.execute("SELECT * FROM receipts WHERE receipt_no=?", (rn,))
      r = cur.fetchone()
  except Exception:
    pass

  cur.execute("SELECT * FROM students WHERE student_id=?", (r["student_id"],))
  s = cur.fetchone()
  cur.execute("SELECT * FROM fee_profile WHERE student_id=?", (r["student_id"],))
  fp = cur.fetchone()
  total, paid, due, status = fee_totals(fp)
  items = _parse_receipt_items(r["fee_for"] or "")
  # If old style, show it as a single line item with receipt amount
  if items and items[0].get("amount") is None:
    items = [{"head": items[0].get("head") or "Fee Payment", "amount": int(r["amount"] or 0), "label": None}]

  paid_sums = payments_by_head(r["student_id"]) if fp else {h:0 for h in FEE_HEADS}
  dues_by_head = headwise_due(fp, paid_sums) if fp else {}
  # Public links for WhatsApp / sharing
  base = request.url_root.rstrip("/")
  tok = (r.get("access_token") if isinstance(r, dict) else r["access_token"])
  seq = (r.get("seq") if isinstance(r, dict) else r["seq"])
  view_link = f"{base}/r/{seq}?t={tok}"
  pdf_link  = f"{base}/r/{seq}/pdf?t={tok}"

  # QR image (data URI). If qrcode isn't installed, silently skip QR.
  qr_uri = ""
  try:
    qr_uri = qr_data_uri(pdf_link)
  except Exception:
    qr_uri = ""

  c.close()
  return render_template(
    "receipt.html",
    college=college_context(),
    receipt=dict(r),
    student=dict(s),
    fee=dict(fp),
    calc={"total": total, "paid": paid, "due": due, "status": status},
    items=items,
    dues_by_head=dues_by_head,
    receipt_year=(r["academic_year"] or get_current_academic_year()),
    view_link=view_link,
    pdf_link=pdf_link,
    qr_uri=qr_uri,
  )

@app.route("/receipt/<path:receipt_no>/regenerate_token", methods=["POST"])
@login_required
def receipt_regenerate_token(receipt_no):
  rn = receipt_no
  try:
    c = conn()
    cur = c.cursor()
    new_tok = secrets.token_urlsafe(16)
    issued = datetime.datetime.now().isoformat(timespec="seconds")
    cur.execute("UPDATE receipts SET access_token=?, token_issued_at=? WHERE receipt_no=?", (new_tok, issued, rn))
    c.commit()
    c.close()
    flash("Receipt link regenerated. Old links will stop working.", "success")
  except Exception:
    try:
      c.close()
    except Exception:
      pass
    flash("Failed to regenerate receipt link.", "error")
  return redirect(url_for("receipt_view", receipt_no=rn))


# Short receipt links for WhatsApp (avoids slash-heavy URLs that can break auto-linking)
@app.route("/r/<int:seq>")
def receipt_short(seq: int):
  """Public-friendly receipt preview page (works with token for WhatsApp)."""
  token = request.args.get("t") or request.args.get("token") or ""
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT * FROM receipts WHERE seq=?", (seq,))
  r = cur.fetchone()
  if not r:
    c.close()
    abort(404)

  # Ensure access_token exists
  if not (r.get("access_token") if isinstance(r, dict) else r["access_token"]):
    try:
      new_tok = secrets.token_urlsafe(16)
      issued = datetime.datetime.now().isoformat(timespec="seconds")
      cur.execute("UPDATE receipts SET access_token=?, token_issued_at=? WHERE seq=?", (new_tok, issued, seq))
      c.commit()
      cur.execute("SELECT * FROM receipts WHERE seq=?", (seq,))
      r = cur.fetchone()
      token = new_tok
    except Exception:
      pass

  # If no session, enforce token
  if not session.get("user"):
    if not _is_valid_public_receipt_token_seq(seq, token):
      c.close()
      abort(403)

  cur.execute("SELECT * FROM students WHERE student_id=?", (r["student_id"],))
  s = cur.fetchone()
  c.close()

  base = request.host_url.rstrip("/")
  view_link = f"{base}/r/{seq}?t={r['access_token']}"
  pdf_link = f"{base}/r/{seq}/pdf?t={r['access_token']}"
  qr_uri = qr_data_uri(pdf_link)

  return render_template(
    "receipt_public.html",
    college=COLLEGE,
    receipt=r,
    student=s,
    view_link=view_link,
    pdf_link=pdf_link,
    qr_uri=qr_uri,
  )


@app.route("/r/<int:seq>/pdf")
def receipt_short_pdf(seq: int):
  # Public access via token (QR/WhatsApp). If not logged in, require valid token.
  if not session.get("user"):
    token = request.args.get("t") or request.args.get("token") or ""
    if not _is_valid_public_receipt_token_seq(seq, token):
      abort(403)

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT receipt_no FROM receipts WHERE seq=?", (seq,))
  row = cur.fetchone()
  c.close()
  if not row:
    abort(404)
  return redirect(url_for('receipt_pdf', receipt_no=row['receipt_no'], t=(request.args.get('t') or request.args.get('token') or '')))


@app.route("/receipt/<path:receipt_no>/pdf")
def receipt_pdf(receipt_no):
  # Public access via token (QR/WhatsApp). If not logged in, require valid token.
  if not session.get("user"):
    token = request.args.get("t") or request.args.get("token") or ""
    if not _is_valid_public_receipt_token_rno(receipt_no, token):
      abort(403)

  rn = receipt_no
  c = conn()
  cur = c.cursor()
  cur.execute("SELECT * FROM receipts WHERE receipt_no=?", (rn,))
  r = cur.fetchone()
  if not r:
    c.close()
    abort(404)
  cur.execute("SELECT * FROM students WHERE student_id=?", (r["student_id"],))
  s = cur.fetchone()
  cur.execute("SELECT * FROM fee_profile WHERE student_id=?", (r["student_id"],))
  fp = cur.fetchone()
  total, paid, due, status = fee_totals(fp)
  items = _parse_receipt_items(r["fee_for"] or "")
  if items and items[0].get("amount") is None:
    items = [{"head": items[0].get("head") or "Fee Payment", "amount": int(r["amount"] or 0), "label": None}]
  paid_sums = payments_by_head(r["student_id"]) if fp else {h:0 for h in FEE_HEADS}
  dues_by_head = headwise_due(fp, paid_sums) if fp else {}
  c.close()

  buf = BytesIO()
  ensure_pdf_fonts()
  w, h = A4
  p = canvas.Canvas(buf, pagesize=A4)

  left = 18*mm
  right = w - 18*mm
  top = h - 18*mm
  y = top

  p.setLineWidth(1)
  p.rect(12*mm, 12*mm, w-24*mm, h-24*mm)

  logo_path = os.path.join(APP_DIR, "static", "images", "kaveri_university_logo.png")
  if os.path.exists(logo_path):
    try:
      img = ImageReader(logo_path)
      p.drawImage(img, left, y-22*mm, width=55*mm, height=18*mm, mask='auto')
    except Exception:
      pass

  pdf_set_font(p, bold=True, size=14)
  p.drawString(left + 60*mm, y-8*mm, COLLEGE["name"])
  pdf_set_font(p, bold=False, size=10)
  p.drawString(left + 60*mm, y-14*mm, COLLEGE["address"])

  pdf_set_font(p, bold=True, size=11)
  p.drawString(left, y-30*mm, f"FEE RECEIPT • Academic Year {COLLEGE.get('academic_year','2025-26')}")

  pdf_set_font(p, bold=False, size=10)
  p.drawString(left, y-40*mm, f"Receipt Number: {r['receipt_no']}")
  p.drawRightString(right, y-40*mm, f"Date: {r['date']}")

  y2 = y - 54*mm
  line_gap = 6.5*mm

  def row(label, value, yy):
    pdf_set_font(p, bold=True, size=10); p.drawString(left, yy, label)
    pdf_set_font(p, bold=False, size=10); p.drawString(left + 45*mm, yy, str(value) if value not in (None,"") else "—")

  row("Student ID", s["student_id"], y2 - 1*line_gap)
  row("Admission Number", s["admission_number"], y2 - 2*line_gap)
  row("Student Name", s["name"], y2 - 3*line_gap)
  row("Course / Branch", f"{s['course']} / {s['branch']}", y2 - 4*line_gap)
  row("Year", s["year"], y2 - 5*line_gap)

  y3 = y2 - 7*line_gap
  pdf_set_font(p, bold=True, size=10); p.drawString(left, y3, "Payment Details")
  pdf_set_font(p, bold=False, size=10)
  p.drawString(left, y3 - 1*line_gap, f"Payment Mode: {r['mode']}")
  p.drawString(left, y3 - 2*line_gap, f"Transaction ID / UTR No: {r['transaction_id'] or '—'}")
  paid_for_labels = []
  for it in items:
    if it.get("head") == "Other Fee" and it.get("label"):
      paid_for_labels.append(it.get("label"))
    else:
      paid_for_labels.append(it.get("head"))
  p.drawString(left, y3 - 3*line_gap, f"Fee Paid For: {', '.join([x for x in paid_for_labels if x])}")

  table_top = y3 - 4.8*line_gap
  pdf_set_font(p, bold=True, size=10); p.drawString(left, table_top, "Paid Fee Heads (This Receipt)")

  tx = left
  ty = table_top - 5*mm
  tw = right - left
  row_h = 9*mm
  col1 = tx + 8*mm
  col2 = tx + tw - 8*mm

  p.setLineWidth(1)
  p.rect(tx, ty - row_h, tw, row_h)
  pdf_set_font(p, bold=True, size=10)
  p.drawString(col1, ty - 6.5*mm, "Fee Head")
  p.drawRightString(col2, ty - 6.5*mm, "Amount (₹)")

  yrow = ty - row_h
  pdf_set_font(p, bold=False, size=10)
  for it in items:
    yrow -= row_h
    p.rect(tx, yrow, tw, row_h)
    head = it.get("head") or "Fee"
    label = it.get("label")
    name = label if (head == "Other Fee" and label) else head
    amt = int(it.get("amount") or 0)
    p.drawString(col1, yrow + 3*mm, name)
    p.drawRightString(col2, yrow + 3*mm, f"{amt:,}")

  yrow -= row_h
  p.rect(tx, yrow, tw, row_h)
  pdf_set_font(p, bold=True, size=10)
  p.drawString(col1, yrow + 3*mm, "TOTAL PAID (This Receipt)")
  p.drawRightString(col2, yrow + 3*mm, f"{int(r['amount'] or 0):,}")

  ysum = yrow - 14*mm
  pdf_set_font(p, bold=True, size=10); p.drawString(left, ysum, "Summary")
  pdf_set_font(p, bold=False, size=10)
  p.drawString(left, ysum - 1*line_gap, f"Total Course Fee: ₹{int(total):,}")
  p.drawString(left, ysum - 2*line_gap, f"Paid (Including this receipt): ₹{int(paid):,}")
  p.drawString(left, ysum - 3*line_gap, f"Due Amount: ₹{int(due):,}")
  pdf_set_font(p, bold=True, size=10)
  p.drawString(left, ysum - 4*line_gap, f"Overall Status: {status.upper()}")

  # Footer + signatory (keep clear spacing to avoid any text overlap)
  footer_base = 18*mm
  footer_gap = 4.8*mm

  pdf_set_font(p, bold=False, size=9)
  for i, line in enumerate(COLLEGE.get("footer_lines", [])):
    # Draw footer lines starting a bit above the bottom border
    p.drawCentredString(w/2, footer_base + (i*footer_gap), line)

  # Signatory block (right aligned, above footer)
  sign_y = footer_base + (len(COLLEGE.get("footer_lines", []))*footer_gap) + 16*mm

  # Optional signature line
  p.setLineWidth(1)
  p.line(right-55*mm, sign_y + 8*mm, right, sign_y + 8*mm)

  sign_text = COLLEGE.get("signatory_text", "Authorised Signatory (Accounts Section)")
  if "(" in sign_text and sign_text.endswith(")"):
    main = sign_text.split("(", 1)[0].strip()
    sub = "(" + sign_text.split("(", 1)[1]
  else:
    main, sub = sign_text, ""

  pdf_set_font(p, bold=True, size=10)
  p.drawRightString(right, sign_y, main)
  if sub:
    pdf_set_font(p, bold=False, size=9)
    p.drawRightString(right, sign_y - 5.5*mm, sub)

  p.showPage()
  p.save()
  buf.seek(0)
  return send_file(buf, mimetype="application/pdf", as_attachment=True,
                   download_name=f"{rn.replace('/','_')}.pdf")



def _xlsx_response(wb, filename):
  from io import BytesIO
  bio = BytesIO()
  wb.save(bio)
  bio.seek(0)
  return send_file(
    bio,
    as_attachment=True,
    download_name=filename,
    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
  )


@app.route("/export/paid_today")
@login_required
def export_paid_today():
  return export_paid_today_excel()

@app.route("/export/paid_today.xlsx")
@login_required
def export_paid_today_excel():
  branch = (request.args.get("branch", "ALL") or "ALL").strip()
  section = (request.args.get("section", "ALL") or "ALL").strip()
  today = datetime.date.today().isoformat()

  c = conn()
  cur = c.cursor()

  where = ""
  params = [today]
  if branch != "ALL":
    where += " AND s.branch=?"
    params.append(branch)
  if section != "ALL":
    where += " AND s.section=?"
    params.append(section)

  cur.execute(
    "SELECT r.receipt_no, r.date, r.fee_for, r.amount, r.mode, r.transaction_id, "
    "s.student_id, s.admission_number, s.name, s.branch, s.section, s.year, s.semester, s.whatsapp "
    "FROM receipts r JOIN students s ON s.student_id=r.student_id "
    "WHERE r.date=?" + where + " AND (s.status IS NULL OR s.status='ACTIVE') "
    "ORDER BY r.seq DESC",
    params,
  )
  rows = cur.fetchall()
  c.close()

  wb = Workbook()
  ws = wb.active
  ws.title = "Paid Today"
  ws.append([
    "Receipt No",
    "Date",
    "Student ID",
    "Admission No",
    "Name",
    "Branch",
    "Section",
    "Year",
    "Semester",
    "Whatsapp",
    "Fee Paid For",
    "Amount",
    "Mode",
    "Transaction/UTR",
  ])

  for r in rows:
    # sqlite3.Row supports mapping access (r["col"]) but not .get()
    fee_for_raw = r["fee_for"] if (hasattr(r, "keys") and ("fee_for" in r.keys())) else None
    items = _parse_receipt_items(fee_for_raw or "")
    labels = []
    if items and items[0].get("amount") is None:
      labels = [items[0].get("head")]
    else:
      for it in items:
        if it.get("head") == "Other Fee" and it.get("label"):
          labels.append(it.get("label"))
        else:
          labels.append(it.get("head"))
    ws.append([
      r["receipt_no"],
      r["date"],
      r["student_id"],
      r["admission_number"],
      r["name"],
      r["branch"],
      r["section"],
      r["year"],
      r["semester"],
      r["whatsapp"],
      ", ".join([x for x in labels if x]),
      int(r["amount"] or 0),
      r["mode"],
      r["transaction_id"],
    ])

  fname = f"Paid_Today_{today}.xlsx"
  return _xlsx_response(wb, fname)


@app.route("/export/not_paid_today")
@login_required
def export_not_paid_today():
  return export_not_paid_today_excel()

@app.route("/export/not_paid_today.xlsx")
@login_required
def export_not_paid_today_excel():
  branch = (request.args.get("branch", "ALL") or "ALL").strip()
  section = (request.args.get("section", "ALL") or "ALL").strip()
  today = datetime.date.today().isoformat()

  c = conn()
  cur = c.cursor()

  w = "WHERE (s.status IS NULL OR s.status='ACTIVE')"
  params = []
  if branch != "ALL":
    w += " AND s.branch=?"
    params.append(branch)
  if section != "ALL":
    w += " AND s.section=?"
    params.append(section)

  # students with no receipt today
  cur.execute(
    "SELECT s.student_id, s.admission_number, s.name, s.branch, s.section, s.year, s.semester, s.whatsapp, "
    "(COALESCE(fp.tuition_fee,0)+COALESCE(fp.hostel_fee,0)+COALESCE(fp.bus_fee,0)+COALESCE(fp.exam_fee,0)+COALESCE(fp.uniform_fee,0)+COALESCE(fp.library_fee,0)+COALESCE(fp.sports_fee,0)+COALESCE(fp.lab_fee,0)+COALESCE(fp.admission_fee,0)+COALESCE(fp.fine_fee,0)+COALESCE(fp.other_fee,0)) AS total_fee, COALESCE(fp.paid_amount,0) AS paid_amount "
    "FROM students s LEFT JOIN fee_profile fp ON fp.student_id=s.student_id "
    + w +
    " AND s.student_id NOT IN (SELECT DISTINCT student_id FROM receipts WHERE date=?) "
    "ORDER BY s.branch, s.section, s.year, s.name",
    params + [today],
  )
  rows = cur.fetchall()
  c.close()

  wb = Workbook()
  ws = wb.active
  ws.title = "Not Paid Today"
  ws.append([
    "Student ID",
    "Admission No",
    "Name",
    "Branch",
    "Section",
    "Year",
    "Semester",
    "Whatsapp",
    "Total Fee",
    "Paid Amount",
    "Due Amount",
  ])

  for r in rows:
    total_fee = int(r["total_fee"] or 0)
    paid_amount = int(r["paid_amount"] or 0)
    due = max(total_fee - paid_amount, 0)
    ws.append([
      r["student_id"],
      r["admission_number"],
      r["name"],
      r["branch"],
      r["section"],
      r["year"],
      r["semester"],
      r["whatsapp"],
      total_fee,
      paid_amount,
      due,
    ])

  fname = f"Not_Paid_Today_{today}.xlsx"
  return _xlsx_response(wb, fname)


@app.route("/collections-analytics")
@login_required
def collections_analytics_page():
  # Filters: today, yesterday, last7, this_month, custom (start/end)
  preset = (request.args.get("range") or "last7").lower()
  today = datetime.date.today()
  if preset == "today":
    start = end = today
  elif preset == "yesterday":
    start = end = today - datetime.timedelta(days=1)
  elif preset == "this_month":
    start = today.replace(day=1)
    end = today
  elif preset == "last7":
    start = today - datetime.timedelta(days=6)
    end = today
  elif preset == "custom":
    try:
      start = datetime.date.fromisoformat((request.args.get("start") or "").strip())
      end = datetime.date.fromisoformat((request.args.get("end") or "").strip())
      if end < start:
        start, end = end, start
    except Exception:
      start = today - datetime.timedelta(days=6)
      end = today
      preset = "last7"
  else:
    start = today - datetime.timedelta(days=6)
    end = today
    preset = "last7"

  start_s = start.isoformat()
  end_s = end.isoformat()
  ay = get_current_academic_year()

  c = conn()
  cur = c.cursor()

  # Total students for current academic year (ACTIVE)
  cur.execute("""SELECT COUNT(*) AS n
                 FROM students
                 WHERE (status IS NULL OR status='ACTIVE')
                   AND (academic_year IS NULL OR academic_year='' OR academic_year=?)""", (ay,))
  total_students = int(cur.fetchone()["n"] or 0)

  # Aggregates for selected date range (receipts/payments)
  cur.execute("""SELECT
      COALESCE(SUM(amount),0) AS total_amount,
      COUNT(*) AS txn_count,
      COUNT(DISTINCT student_id) AS paid_students
    FROM receipts
    WHERE date BETWEEN ? AND ?
      AND (academic_year IS NULL OR academic_year='' OR academic_year=?)""", (start_s, end_s, ay))
  agg = cur.fetchone()
  total_amount = int(agg["total_amount"] or 0)
  txn_count = int(agg["txn_count"] or 0)
  paid_students = int(agg["paid_students"] or 0)
  not_paid = max(total_students - paid_students, 0)

  # Daily chart (within selected range)
  cur.execute("""SELECT date, COALESCE(SUM(amount),0) AS amt
                 FROM receipts
                 WHERE date BETWEEN ? AND ?
                   AND (academic_year IS NULL OR academic_year='' OR academic_year=?)
                 GROUP BY date
                 ORDER BY date ASC""", (start_s, end_s, ay))
  daily_rows = cur.fetchall()
  daily_labels = [r["date"] for r in daily_rows]
  daily_values = [int(r["amt"] or 0) for r in daily_rows]

  # Monthly chart (last 6 months, calendar)
  months = []
  m0 = today.replace(day=1)
  for i in range(5, -1, -1):
    mm = (m0.month - i)
    yy = m0.year
    while mm <= 0:
      mm += 12
      yy -= 1
    months.append(f"{yy:04d}-{mm:02d}")
  placeholders = ",".join(["?"]*len(months))
  cur.execute(f"""SELECT substr(date,1,7) AS ym, COALESCE(SUM(amount),0) AS amt
                  FROM receipts
                  WHERE substr(date,1,7) IN ({placeholders})
                    AND (academic_year IS NULL OR academic_year='' OR academic_year=?)
                  GROUP BY ym""", (*months, ay))
  mm_rows = {r["ym"]: int(r["amt"] or 0) for r in cur.fetchall()}
  monthly_labels = months
  monthly_values = [mm_rows.get(m, 0) for m in months]

  # Recent transactions (top 50)
  cur.execute("""SELECT r.receipt_no, r.date, r.student_id, s.name AS student_name, r.amount, r.mode, r.transaction_id
                 FROM receipts r
                 LEFT JOIN students s ON s.student_id = r.student_id
                 WHERE r.date BETWEEN ? AND ?
                   AND (r.academic_year IS NULL OR r.academic_year='' OR r.academic_year=?)
                 ORDER BY r.date DESC, r.seq DESC
                 LIMIT 50""", (start_s, end_s, ay))
  tx = [dict(row) for row in cur.fetchall()]
  c.close()

  return render_template(
    "collections_analytics.html",
    college=college_context(),
    preset=preset,
    start=start_s,
    end=end_s,
    total_amount=total_amount,
    txn_count=txn_count,
    paid_students=paid_students,
    not_paid=not_paid,
    total_students=total_students,
    daily_labels=json.dumps(daily_labels),
    daily_values=json.dumps(daily_values),
    monthly_labels=json.dumps(monthly_labels),
    monthly_values=json.dumps(monthly_values),
    tx=tx
  )

@app.route("/reports")
@login_required
def reports_page():
  c = conn()
  cur = c.cursor()
  cur.execute("""SELECT s.student_id, s.name, s.branch, s.year, fp.*
                 FROM students s JOIN fee_profile fp ON fp.student_id=s.student_id
                 ORDER BY s.student_id""")
  rows = []
  for r in cur.fetchall():
    total, paid, due, status = fee_totals(r)
    d = dict(r)
    d["total"] = total
    d["due"] = due
    d["status"] = status
    rows.append(d)
  c.close()
  due_students = [x for x in rows if x["due"] > 0]
  paid_students = [x for x in rows if x["due"] == 0]
  return render_template("reports.html", college=college_context(), rows=rows, due_students=due_students, paid_students=paid_students)


@app.route("/export/paid_range.xlsx")
@login_required
def export_paid_range_xlsx():
  # Query params: start=YYYY-MM-DD, end=YYYY-MM-DD (inclusive)
  start = (request.args.get("start") or "").strip()
  end = (request.args.get("end") or "").strip()
  if not start or not end:
    abort(400)
  try:
    sdate = datetime.date.fromisoformat(start)
    edate = datetime.date.fromisoformat(end)
    if edate < sdate:
      sdate, edate = edate, sdate
  except Exception:
    abort(400)
  start = sdate.isoformat()
  end = edate.isoformat()
  ay = get_current_academic_year()

  c = conn()
  cur = c.cursor()
  cur.execute("""SELECT r.date, r.receipt_no, r.student_id, s.name AS student_name, r.amount, r.mode, r.transaction_id
                 FROM receipts r
                 LEFT JOIN students s ON s.student_id = r.student_id
                 WHERE r.date BETWEEN ? AND ?
                   AND (r.academic_year IS NULL OR r.academic_year='' OR r.academic_year=?)
                 ORDER BY r.date ASC, r.seq ASC""", (start, end, ay))
  rows = [dict(r) for r in cur.fetchall()]
  c.close()

  wb = Workbook()
  ws = wb.active
  ws.title = "Paid"
  ws.append(["Date","Receipt No","Student ID","Student Name","Amount","Mode","Transaction/UTR"])
  for r in rows:
    ws.append([
      r.get("date"),
      r.get("receipt_no"),
      r.get("student_id"),
      r.get("student_name"),
      int(r.get("amount") or 0),
      r.get("mode"),
      r.get("transaction_id"),
    ])
  fname = f"Paid_{start}_to_{end}.xlsx"
  return _xlsx_response(wb, fname)


@app.route("/export/not_paid_range.xlsx")
@login_required
def export_not_paid_range_xlsx():
  # Not paid = ACTIVE students with no receipt in selected date range.
  # IMPORTANT: Do not depend on academic_year formatting (prevents "only first 4" bug).
  start = (request.args.get("start") or "").strip()
  end = (request.args.get("end") or "").strip()
  if not start or not end:
    abort(400)
  try:
    sdate = datetime.date.fromisoformat(start)
    edate = datetime.date.fromisoformat(end)
    if edate < sdate:
      sdate, edate = edate, sdate
  except Exception:
    abort(400)
  start = sdate.isoformat()
  end = edate.isoformat()
  c = conn()
  cur = c.cursor()
  cur.execute("""SELECT s.student_id, s.admission_number, s.name, s.course, s.branch, s.year, s.semester, s.section, s.whatsapp, s.email,
                      fp.tuition_fee, fp.hostel_fee, fp.bus_fee, fp.exam_fee, fp.uniform_fee, fp.library_fee, fp.sports_fee, fp.lab_fee,
                      fp.admission_fee, fp.fine_fee, fp.other_fee
                 FROM students s
                 LEFT JOIN fee_profile fp ON fp.student_id = s.student_id
                 WHERE (s.status IS NULL OR s.status='ACTIVE')
                   AND s.student_id NOT IN (
                     SELECT DISTINCT student_id
                     FROM receipts
                     WHERE date BETWEEN ? AND ?
                   )
                 ORDER BY s.branch ASC, s.section ASC, s.student_id ASC""", (start, end))
  students_rows = [dict(r) for r in cur.fetchall()]

  # Aggregate all-time paid amounts per student per head (single pass) for head-wise due.
  sids = [r["student_id"] for r in students_rows]
  paid_map = {sid: {h: 0 for h in FEE_HEADS} for sid in sids}
  if sids:
    qmarks = ",".join(["?"] * len(sids))
    cur.execute(f"SELECT student_id, fee_for, amount FROM receipts WHERE student_id IN ({qmarks})", sids)
    for rr in cur.fetchall():
      sid = rr["student_id"]
      if sid not in paid_map:
        continue
      items = _parse_receipt_items(rr["fee_for"] or "")
      if items and all(it.get("amount") is not None for it in items):
        for it in items:
          head = it.get("head")
          if head in paid_map[sid]:
            try:
              paid_map[sid][head] += int(it.get("amount") or 0)
            except Exception:
              pass
      else:
        # old receipt fallback
        try:
          amt = int(rr["amount"] or 0)
        except Exception:
          amt = 0
        # Put into Other Fee bucket
        paid_map[sid]["Other Fee"] += amt

  c.close()

  wb = Workbook()
  ws = wb.active
  ws.title = "Not Paid"

  head_cols = FEE_HEADS[:]  # exact column names
  ws.append([
    "Student ID","Admission No","Name","Course","Branch","Year","Semester","Section","WhatsApp","Email",
    *head_cols,
    "Total Balance Due"
  ])

  for r in students_rows:
    sid = r.get("student_id")
    planned = {
      "Tuition Fee": int(r.get("tuition_fee") or 0),
      "Hostel Fee": int(r.get("hostel_fee") or 0),
      "Bus Fee": int(r.get("bus_fee") or 0),
      "Exam Fee": int(r.get("exam_fee") or 0),
      "Uniform Fee": int(r.get("uniform_fee") or 0),
      "Library Fee": int(r.get("library_fee") or 0),
      "Sports Fee": int(r.get("sports_fee") or 0),
      "Lab Fee": int(r.get("lab_fee") or 0),
      "Admission Fee": int(r.get("admission_fee") or 0),
      "Fine": int(r.get("fine_fee") or 0),
      "Other Fee": int(r.get("other_fee") or 0),
    }
    paid_sums = paid_map.get(sid, {h: 0 for h in FEE_HEADS})
    due_vals = []
    total_due = 0
    for h in head_cols:
      due_h = max(int(planned.get(h, 0)) - int(paid_sums.get(h, 0)), 0)
      due_vals.append(due_h)
      total_due += due_h

    ws.append([
      sid,
      r.get("admission_number"),
      r.get("name"),
      r.get("course"),
      r.get("branch"),
      r.get("year"),
      r.get("semester"),
      r.get("section"),
      r.get("whatsapp"),
      r.get("email"),
      *due_vals,
      total_due
    ])
  fname = f"NotPaid_{start}_to_{end}.xlsx"
  return _xlsx_response(wb, fname)

@app.route("/api/backup", methods=["POST"])
@login_required
@admin_required
def api_backup():
  os.makedirs(BACKUP_DIR, exist_ok=True)
  stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
  dst = os.path.join(BACKUP_DIR, f"database_backup_{stamp}.db")
  if not os.path.exists(DB_PATH):
    init_db()
  shutil.copy2(DB_PATH, dst)
  return jsonify({"ok": True, "file": os.path.basename(dst)})

@app.route("/backup/<filename>")
@login_required
@admin_required
def download_backup(filename):
  path = os.path.join(BACKUP_DIR, filename)
  if not os.path.exists(path):
    abort(404)
  return send_file(path, as_attachment=True)

# =====================
# STUDENTS + FEES CSV (Excel) IMPORT / EXPORT
# =====================
STUDENT_CSV_COLUMNS = [
  "student_id","admission_number","name","course","branch","year","semester","section","academic_year","whatsapp","email",
  "Tuition Fee","Hostel Fee","Bus Fee","Exam Fee","Uniform Fee","Library Fee","Sports Fee","Lab Fee","Admission Fee","Fine","Other Fee"
]

@app.route("/export/students_full.csv")
@login_required
def export_students_full_csv():
  """Export students (optionally folder filtered).

  Query params:
    view=active|removed|unassigned
    branch=...
    section=...
  """
  view = (request.args.get("view", "active") or "active").lower().strip()
  branch = (request.args.get("branch", "") or "").strip()
  section = (request.args.get("section", "") or "").strip()

  where = []
  params = []
  if view == "removed":
    where.append("IFNULL(s.status,'ACTIVE')='REMOVED'")
  else:
    where.append("IFNULL(s.status,'ACTIVE')='ACTIVE'")

  if view == "unassigned":
    where.append("(s.branch IS NULL OR TRIM(s.branch)='')")
  else:
    if branch:
      where.append("UPPER(TRIM(s.branch))=?")
      params.append(branch.upper())
    if section:
      where.append("UPPER(TRIM(s.section))=?")
      params.append(section.upper())

  where_sql = "WHERE " + " AND ".join(where) if where else ""

  c = conn()
  cur = c.cursor()
  cur.execute(f"""SELECT s.student_id,s.admission_number,s.name,s.course,s.branch,s.year,s.semester,s.section,s.academic_year,s.whatsapp,s.email,
                    fp.tuition_fee,fp.hostel_fee,fp.bus_fee,fp.exam_fee,fp.uniform_fee,fp.library_fee,fp.sports_fee,fp.lab_fee,
                    fp.admission_fee,fp.fine_fee,fp.other_fee
                 FROM students s
                 LEFT JOIN fee_profile fp ON fp.student_id=s.student_id
                 {where_sql}
                 ORDER BY s.student_id""", tuple(params))
  rows = [dict(r) for r in cur.fetchall()]
  c.close()

  out = StringIO()
  w = csv.writer(out)
  w.writerow(STUDENT_CSV_COLUMNS)
  for r in rows:
    out_row = []
    for col in STUDENT_CSV_COLUMNS:
      if col in FEE_PROFILE_COLS:
        out_row.append(r.get(FEE_PROFILE_COLS[col], 0))
      else:
        out_row.append(r.get(col))
    w.writerow(out_row)

  data = out.getvalue().encode("utf-8")
  # Name includes folder context if selected
  suffix = "all"
  if view == "removed":
    suffix = "removed"
  elif view == "unassigned":
    suffix = "unassigned"
  elif branch and section:
    suffix = f"{branch}_{section}"
  elif branch:
    suffix = f"{branch}"

  fname = f"students_{suffix}_{COLLEGE.get('academic_year','2025-26')}.csv"
  return send_file(BytesIO(data), mimetype="text/csv", as_attachment=True, download_name=fname)

@app.route("/export/students_sample.csv")
@login_required
def export_students_sample_csv():
  out = StringIO()
  w = csv.writer(out)
  w.writerow(STUDENT_CSV_COLUMNS)
  w.writerow([
    "2502UG0201","ADM-2025-0201","STUDENT NAME","B.Tech","CSE","1","1","A",COLLEGE.get("academic_year","2025-26"),"9999999999","student@email.com",
    "185000","0","0","0","4000","0","0","0","0","0","0"
  ])
  data = out.getvalue().encode("utf-8")
  return send_file(BytesIO(data), mimetype="text/csv", as_attachment=True, download_name="students_sample.csv")

@app.route("/import/students_full", methods=["POST"])
@login_required
@admin_required
def import_students_full_csv():
  # Optional folder context: if a folder is selected on Students page, we
  # can force-import into that branch/section.
  force_branch = (request.form.get("branch") or "").strip()
  force_section = (request.form.get("section") or "").strip()
  force_view = (request.form.get("view") or "active").strip().lower()

  if "file" not in request.files:
    return jsonify({"ok": False, "error": "CSV file required"}), 400

  # Read bytes first so we can detect common wrong uploads (like .xlsx).
  data_bytes = request.files["file"].read()
  if not data_bytes:
    return jsonify({"ok": False, "error": "Uploaded file is empty"}), 400

  # XLSX files are ZIPs and typically start with 'PK'.
  if data_bytes[:2] == b"PK":
    return jsonify({
      "ok": False,
      "error": "Please upload a CSV file (not .xlsx). Use: Download Sample CSV → fill → Save as CSV (UTF-8)."
    }), 400

  raw = data_bytes.decode("utf-8-sig", errors="ignore")  # handles Excel BOM
  reader = csv.DictReader(StringIO(raw))

  # validate headers
  # Normalize headers by stripping spaces/BOM.
  fieldnames = [f.strip() for f in (reader.fieldnames or [])]
  missing = [c for c in STUDENT_CSV_COLUMNS if c not in fieldnames]
  if missing:
    return jsonify({"ok": False, "error": "Missing columns: " + ", ".join(missing)}), 400

  c = conn()
  cur = c.cursor()
  added = updated = skipped = 0

  for row in reader:
    # Normalize header keys (helps if someone adds spaces in Excel headers).
    row = { (k.strip() if isinstance(k, str) else k): v for k, v in (row or {}).items() }

    sid = (row.get("student_id") or "").strip().upper()
    name = (row.get("name") or "").strip()
    if not sid or not name:
      skipped += 1
      continue

    # student fields
    # Apply folder context if provided (import section-wise)
    if force_view == "unassigned":
      row["branch"] = ""
      row["section"] = ""
    else:
      if force_branch:
        row["branch"] = force_branch
      if force_section:
        row["section"] = force_section

    svals = (
      (row.get("admission_number") or "").strip(),
      name,
      (row.get("course") or "").strip(),
      (row.get("branch") or "").strip(),
      (row.get("year") or "").strip(),
      (row.get("semester") or "").strip(),
      (row.get("section") or "").strip(),
      (row.get("academic_year") or "").strip(),
      (row.get("whatsapp") or "").strip(),
      (row.get("email") or "").strip(),
    )

    cur.execute("SELECT 1 FROM students WHERE student_id=?", (sid,))
    exists = cur.fetchone() is not None
    if not exists:
      cur.execute("""INSERT INTO students(student_id, admission_number, name, course, branch, year, semester, section, academic_year, whatsapp, email, status)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", (sid, *svals, 'ACTIVE'))
      cur.execute("INSERT OR IGNORE INTO fee_profile(student_id) VALUES(?)", (sid,))
      added += 1
    else:
      cur.execute("""UPDATE students SET admission_number=?, name=?, course=?, branch=?, year=?, semester=?, section=?, academic_year=?, whatsapp=?, email=?
                    WHERE student_id=?""", (*svals, sid))
      updated += 1

    # If the student was previously removed, re-activate on import.
    cur.execute("UPDATE students SET status='ACTIVE' WHERE student_id=?", (sid,))

    # fee fields
    def to_int(x):
      try:
        return int(str(x).strip() or "0")
      except Exception:
        return 0
    fvals = (
      to_int(row.get("Tuition Fee")),
      to_int(row.get("Hostel Fee")),
      to_int(row.get("Bus Fee")),
      to_int(row.get("Exam Fee")),
      to_int(row.get("Uniform Fee")),
      to_int(row.get("Library Fee")),
      to_int(row.get("Sports Fee")),
      to_int(row.get("Lab Fee")),
      to_int(row.get("Admission Fee")),
      to_int(row.get("Fine")),
      to_int(row.get("Other Fee")),
      sid
    )
    cur.execute("""UPDATE fee_profile SET tuition_fee=?, hostel_fee=?, bus_fee=?, exam_fee=?, uniform_fee=?, library_fee=?, sports_fee=?, lab_fee=?,
                                      admission_fee=?, fine_fee=?, other_fee=?
                 WHERE student_id=?""", fvals)

  c.commit()
  c.close()
  return jsonify({"ok": True, "added": added, "updated": updated, "skipped": skipped})



@app.route("/settings")
@login_required
def settings_page():
  # Admin can manage users; all users can change their own password.
  c = conn()
  cur = c.cursor()
  users = []
  try:
    cur.execute("SELECT username, role, created_at FROM users ORDER BY role DESC, username ASC")
    users = [dict(r) for r in cur.fetchall()]
  except Exception:
    users = []
  c.close()
  return render_template("settings.html", college=college_context(), users=users, user=session.get("user"), role=session.get("role"))


@app.route("/settings/academic_year", methods=["POST"])
@login_required
@admin_required
def settings_academic_year():
  ay = (request.form.get("academic_year") or "").strip()
  if not ay:
    flash("Academic Year is required.", "danger")
    return redirect(url_for("settings_page"))
  set_setting("current_academic_year", ay)
  flash("Academic Year updated successfully.", "success")
  return redirect(url_for("settings_page"))

@app.route("/settings/change_password", methods=["POST"])
@login_required
def settings_change_password():
  current = (request.form.get("current_password") or "").strip()
  new1 = (request.form.get("new_password") or "").strip()
  new2 = (request.form.get("confirm_password") or "").strip()
  if not current or not new1 or not new2:
    flash("Please fill all password fields.", "danger")
    return redirect(url_for("settings_page"))
  if new1 != new2:
    flash("New password and confirm password do not match.", "danger")
    return redirect(url_for("settings_page"))

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT password_hash FROM users WHERE username=?", (session.get("user"),))
  row = cur.fetchone()
  if not row or not check_password_hash(row["password_hash"], current):
    c.close()
    flash("Current password is incorrect.", "danger")
    return redirect(url_for("settings_page"))

  cur.execute("UPDATE users SET password_hash=? WHERE username=?", (generate_password_hash(new1), session.get("user")))
  c.commit()
  c.close()
  flash("Password updated successfully.", "success")
  return redirect(url_for("settings_page"))

@app.route("/settings/users/create", methods=["POST"])
@login_required
@admin_required
def settings_users_create():
  username = (request.form.get("username") or "").strip()
  password = (request.form.get("password") or "").strip()
  role = (request.form.get("role") or "staff").strip().lower()
  if role not in ("admin", "staff"):
    role = "staff"
  if not username or not password:
    flash("Username and password are required.", "danger")
    return redirect(url_for("settings_page"))

  c = conn()
  cur = c.cursor()
  cur.execute("SELECT username FROM users WHERE username=?", (username,))
  if cur.fetchone():
    c.close()
    flash("User already exists.", "danger")
    return redirect(url_for("settings_page"))

  cur.execute(
    "INSERT INTO users(username, password_hash, role, created_at) VALUES(?,?,?,?)",
    (username, generate_password_hash(password), role, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
  )
  c.commit()
  c.close()
  flash(f"User '{username}' created.", "success")
  return redirect(url_for("settings_page"))

@app.route("/settings/users/reset_password", methods=["POST"])
@login_required
@admin_required
def settings_users_reset_password():
  username = (request.form.get("username") or "").strip()
  new_password = (request.form.get("new_password") or "").strip()
  if not username or not new_password:
    flash("Username and new password are required.", "danger")
    return redirect(url_for("settings_page"))
  if username == session.get("user"):
    flash("Use Change Password to update your own password.", "danger")
    return redirect(url_for("settings_page"))

  c = conn()
  cur = c.cursor()
  cur.execute("UPDATE users SET password_hash=? WHERE username=?", (generate_password_hash(new_password), username))
  if cur.rowcount == 0:
    c.close()
    flash("User not found.", "danger")
    return redirect(url_for("settings_page"))
  c.commit()
  c.close()
  flash(f"Password reset for '{username}'.", "success")
  return redirect(url_for("settings_page"))

@app.route("/settings/users/set_role", methods=["POST"])
@login_required
@admin_required
def settings_users_set_role():
  username = (request.form.get("username") or "").strip()
  role = (request.form.get("role") or "").strip().lower()
  if role not in ("admin", "staff"):
    role = "staff"
  if not username:
    flash("Username required.", "danger")
    return redirect(url_for("settings_page"))
  if username == session.get("user") and role != "admin":
    flash("You cannot remove your own admin access.", "danger")
    return redirect(url_for("settings_page"))

  c = conn()
  cur = c.cursor()
  cur.execute("UPDATE users SET role=? WHERE username=?", (role, username))
  if cur.rowcount == 0:
    c.close()
    flash("User not found.", "danger")
    return redirect(url_for("settings_page"))
  c.commit()
  c.close()
  flash(f"Role updated for '{username}'.", "success")
  return redirect(url_for("settings_page"))

@app.route("/settings/users/delete", methods=["POST"])
@login_required
@admin_required
def settings_users_delete():
  username = (request.form.get("username") or "").strip()
  if not username:
    flash("Username required.", "danger")
    return redirect(url_for("settings_page"))
  if username == session.get("user"):
    flash("You cannot delete your own account.", "danger")
    return redirect(url_for("settings_page"))

  c = conn()
  cur = c.cursor()
  cur.execute("DELETE FROM users WHERE username=?", (username,))
  if cur.rowcount == 0:
    c.close()
    flash("User not found.", "danger")
    return redirect(url_for("settings_page"))
  c.commit()
  c.close()
  flash(f"User '{username}' deleted.", "success")
  return redirect(url_for("settings_page"))



if __name__ == "__main__":
  init_db()
  # Run on all interfaces for LAN access
  app.run(host="0.0.0.0", port=5000, debug=True)